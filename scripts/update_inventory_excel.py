import openpyxl

def replace_inventory_data(excel_path, sheet_name):
    data = [
        ['No Item', 'Product Name', 'Stock Remaining', 'Unit Price', 'Total Value', 'Gross Profit'],
        ['ITEM-001', 'Madu Multiflora', 34, 84000, 2856000, 2772000],
        ['ITEM-002', 'Madu Klengkeng', 19, 100000, 1900000, 1800000],
        ['ITEM-003', 'Madu Kapuk Randu', 22, 96000, 2112000, 2016000],
        ['ITEM-004', 'Madu Hutan', 20, 80000, 1600000, 1520000],
        ['ITEM-005', 'Madu Karet', 20, 100000, 2000000, 1900000],
        ['ITEM-006', 'Madu Cengkeh', 32, 96000, 3072000, 2976000],
        ['ITEM-007', 'Madu Kopi', 22, 92000, 2025000, 1933000],
        ['ITEM-008', 'Madu Rambutan', 25, 92000, 2300000, 2208000],
        ['ITEM-009', 'Madu Mangga', 18, 92000, 1656000, 1564000],
        ['ITEM-010', 'Madu Pahitan', 35, 108000, 3780000, 3672000],
        ['ITEM-011', 'Madu Royal Jelly', 28, 280000, 7840000, 7560000],
        ['ITEM-012', 'Honey Comb', 25, 130000, 3600000, 3470000]
    ]

    wb = openpyxl.load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Clear sheet content
        for row in ws['A1':f'F{ws.max_row}']:
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(excel_path)
    print(f"Inventory data inserted into {excel_path}, sheet: {sheet_name}")

if __name__ == "__main__":
    excel_path = 'BeeTheOne/inventory_data.xlsx'
    sheet_name = 'Inventory'
    replace_inventory_data(excel_path, sheet_name)
